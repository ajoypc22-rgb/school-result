"""
excel_to_json.py
-----------------
Converts the school's mark-entry Excel sheet (subject blocks of
মোট নম্বর / প্রাপ্ত নম্বর / গ্রেড [/ পয়েন্ট]) into results.json for the
Result Portal. Works whether or not the sheet has a "Roll" column, and
whether each subject block is 3 or 4 columns wide.

Usage:
    python3 excel_to_json.py <input.xlsx> [output.json]

Re-run this any time after filling in marks in the Excel file to
regenerate results.json. No code changes needed.
"""
import sys
import json
import re
import openpyxl

# ---- grade boundaries, copied exactly from the Excel formulas ---------
# Keyed by a subject's total mark (30 / 25 / 15) — the sheet uses a
# slightly different cut-off table for each, so we replicate them as-is
# instead of a single percentage scale.
BOUNDARIES = {
    30: [(23, "A+"), (21, "A"), (17, "A-"), (14, "B"), (11, "C"), (8.5, "D")],
    25: [(17, "A+"), (14.5, "A"), (12.5, "A-"), (10.5, "B"), (8.5, "C"), (6.5, "D")],
    15: [(11, "A+"), (10, "A"), (8, "A-"), (7, "B"), (6, "C"), (4, "D")],
}


def grade_for(obtained, total):
    if obtained is None:
        obtained = 0
    table = BOUNDARIES.get(total)
    if table is None:
        table = [(t / 30 * total, g) for t, g in BOUNDARIES[30]]
    for cutoff, letter in table:
        if obtained > cutoff:
            return letter
    return "F"


STOP_WORDS = ("জি.পি.এ", "সর্বমোট", "failed", "GPA")


def find_layout(ws):
    """Locate the Roll / name columns and every subject's column block."""
    header_row = 4
    sub_row = 5

    roll_col = None
    name_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        v = str(v).strip()
        if v.lower() in ("roll", "রোল"):
            roll_col = c
        elif v == "নাম":
            name_col = c

    if name_col is None:
        raise ValueError("Could not find a 'নাম' header column on row 4.")

    # collect subject header start columns, in order, after name_col
    subject_starts = []
    for c in range(name_col + 1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        text = str(v).strip()
        if any(sw in text for sw in STOP_WORDS):
            break
        subject_starts.append((c, text))

    subjects = []  # (name, total_col, obtained_col, grade_col)
    for i, (start_col, subj_name) in enumerate(subject_starts):
        end_col = subject_starts[i + 1][0] if i + 1 < len(subject_starts) else ws.max_column + 1
        total_col = obtained_col = grade_col = None
        for c in range(start_col, end_col):
            label = ws.cell(row=sub_row, column=c).value
            if label is None:
                continue
            label = str(label).strip()
            if "মোট" in label:
                total_col = c
            elif "প্রাপ্ত" in label:
                obtained_col = c
            elif "গ্রেড" in label:
                grade_col = c
        if total_col and obtained_col:
            subjects.append((subj_name.strip(), total_col, obtained_col, grade_col))

    return roll_col, name_col, subjects


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 excel_to_json.py <input.xlsx> [output.json]")
        sys.exit(1)

    in_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "results.json"

    wb = openpyxl.load_workbook(in_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    # title can be in A1 or B1 depending on the sheet's layout
    title_cell = ""
    for c in range(1, 4):
        v = ws.cell(row=1, column=c).value
        if v:
            title_cell = str(v)
            break

    exam_name = title_cell.split("ফলাফল")[0].strip()
    class_match = re.search(r"শ্রেণিঃ\s*([^\s]+)", title_cell)
    section_match = re.search(r"শাখাঃ\s*(.+)", title_cell)
    exam_class = (class_match.group(1) if class_match else "").strip()
    section = (section_match.group(1) if section_match else "").strip()
    section = re.sub(r"\s+", " ", section)

    roll_col, name_col, subjects = find_layout(ws)
    if not subjects:
        print("Could not detect subject headers on row 4 — check the sheet layout.")
        sys.exit(1)

    students = []
    auto_roll = 1
    for row in range(6, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()

        if roll_col:
            roll_val = ws.cell(row=row, column=roll_col).value
            roll = str(roll_val).strip() if roll_val not in (None, "") else str(auto_roll)
        else:
            roll = str(auto_roll)

        subj_rows = []
        for subj_name, total_col, obtained_col, grade_col in subjects:
            total = ws.cell(row=row, column=total_col).value
            obtained = ws.cell(row=row, column=obtained_col).value
            total = total if isinstance(total, (int, float)) else 0
            obtained = obtained if isinstance(obtained, (int, float)) else None
            grade = grade_for(obtained, total)
            subj_rows.append({
                "name": subj_name,
                "total": total,
                "obtained": obtained if obtained is not None else 0,
                "grade": grade
            })

        students.append({
            "roll": roll,
            "name": name,
            "class": exam_class,
            "section": section,
            "subjects": subj_rows
        })
        auto_roll += 1

    data = {
        "school": {
            "name": "প্রবর্তক স্কুল এন্ড কলেজ",
            "address": "পাঁচলাইশ, চট্টগ্রাম",
            "logo": "logo.png",
            "examName": exam_name,
            "examClass": f"{exam_class} শ্রেণি" + (f" ({section})" if section else "")
        },
        "students": students
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(students)} students -> {out_path}")
    print(f"Roll column detected: {'yes' if roll_col else 'no (auto-numbered)'}")
    print(f"Subjects detected: {[s[0] for s in subjects]}")


if __name__ == "__main__":
    main()
